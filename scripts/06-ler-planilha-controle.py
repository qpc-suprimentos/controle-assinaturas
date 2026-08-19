# -*- coding: utf-8 -*-
"""
LE A PLANILHA F.SUP.G200.002 E VIRA A BASE CONSOLIDADA DO PAINEL
================================================================

O QUE ESTE SCRIPT FAZ, EM PORTUGUES SIMPLES:

Suas colegas mantem a planilha "Controle de assinatura de contratos", que tem
TODOS os contratos da obra sob supervisao de voces - nao so os que voce recebe
por e-mail. Ela e o retrato completo; os e-mails sao o que acontece depois dele.

Este script le essa planilha e transforma em um arquivo de dados que o gerador
do painel entende. Ele nao inventa nada e nao corrige nada: so traduz.

A planilha original NUNCA e alterada. Ela fica em 01-dados-brutos/ e e so leitura.

COMO RODAR:
    python 05-scripts/06-ler-planilha-controle.py
"""

import json
import os
import re
import unicodedata
from datetime import datetime
from zoneinfo import ZoneInfo

import openpyxl

# =============================================================================
# CONFIGURACAO - e so aqui que voce mexe
# =============================================================================

PASTA_RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PASTA_BRUTOS = os.path.join(PASTA_RAIZ, "01-dados-brutos")
ARQUIVO_SAIDA = os.path.join(PASTA_BRUTOS, "base-planilha.json")

ABA = "Fluxo de assinatura"
LINHA_DATA_ENVIO = 4        # "DATA DE ENVIO: 17/08/2026"
LINHA_CABECALHO_SIG = 6     # nomes dos signatarios
PRIMEIRA_LINHA_DADOS = 7

COL_NUMERO = 1
COL_IDENTIFICACAO = 2
COL_RAZAO_SOCIAL = 3
COL_SERVICO = 4
COL_CNPJ = 5
COL_DATA_CADASTRO = 6
COL_DIAS_SEM_ASSINATURA = 7
COL_PRIMEIRO_SIGNATARIO = 9
COL_ULTIMO_SIGNATARIO = 19

FUSO_LOCAL = ZoneInfo("America/Sao_Paulo")

# Decisao do Valter em 19/08/2026: esta planilha cobre APENAS contratos LYON.
# AFE e QPC tem outro controle, que ainda nao chegou.
TIPO_DESTA_PLANILHA = "LYON"


# =============================================================================
# REGRAS DE LEITURA DAS CELULAS - confirmadas pelo Valter em 19/08/2026
# =============================================================================
#
# O que pode aparecer numa celula de assinatura, e o que cada coisa significa:
#
#   "ASSINADO\n09/02/2026"  -> assinou, nesta data
#   "ASSINADO" (sem data)   -> assinou; a data faltou no preenchimento
#   "PENDENTE"              -> falta assinar
#   "N.A"                   -> NAO PRECISOU passar por essa pessoa
#   "-"                     -> NAO PRECISOU passar por essa pessoa
#   celula vazia            -> falta assinar
#
# O ponto que mais importa: "N.A" e "-" saem da conta inteira. Um contrato onde
# duas pessoas nao precisavam assinar e que tem as outras nove assinadas esta
# 9 de 9, ou seja, COMPLETO - e nao 9 de 11, que pareceria travado para sempre.

PADRAO_ASSINADO = re.compile(r"^\s*ASSINADO", re.IGNORECASE)
PADRAO_DATA = re.compile(r"(\d{2}/\d{2}/\d{4})")
NAO_SE_APLICA = {"N.A", "N.A.", "NA", "-", "--"}


def erro(mensagem):
    """Para tudo e avisa em portugues. Regra do projeto: falhar alto, nao baixo."""
    raise SystemExit("\n[ERRO] " + mensagem + "\n")


def sem_acento(texto):
    return "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )


def achar_planilha():
    """Pega a planilha de controle mais recente da pasta de dados brutos."""
    if not os.path.isdir(PASTA_BRUTOS):
        erro("A pasta 01-dados-brutos nao existe.")
    candidatos = sorted(
        f for f in os.listdir(PASTA_BRUTOS)
        if f.startswith("controle-assinaturas-") and f.endswith(".xlsx")
    )
    if not candidatos:
        erro(
            "Nao achei nenhuma planilha controle-assinaturas-AAAA-MM-DD.xlsx em\n"
            "01-dados-brutos/. Salve ali a planilha F.SUP.G200.002 que suas colegas mantem."
        )
    return os.path.join(PASTA_BRUTOS, candidatos[-1])


def chave_do_documento(identificacao):
    """
    Cria a etiqueta que liga a linha da planilha ao documento do e-mail.

    A planilha escreve de dois jeitos ("CT-LSFG200-063.26" e "CT-LSF-G200-063.26")
    e o e-mail escreve de um terceiro ("63.CT-LSF-G200-063-26 - INOSERVICE.pdf").
    Todos viram a mesma etiqueta: o que importa e ser aditivo ou nao, e o numero.

      "CT-LSF-G200-063.26"          -> CT-063
      "CT-LSFG200-063.26"           -> CT-063
      "ADIT.V.01-LSF-G200-046.26"   -> ADIT01-046
    """
    texto = sem_acento(str(identificacao)).upper()

    aditivo = re.search(r"ADIT\.?\s*V?\.?\s*(\d{1,2})", texto)
    numero = re.search(r"G200[-\s.]*(\d{2,3})", texto)
    if not numero:
        return None, False

    if aditivo:
        return "ADIT%s-%s" % (aditivo.group(1).zfill(2), numero.group(1).zfill(3)), True
    return "CT-%s" % numero.group(1).zfill(3), False


def ler_celula_assinatura(valor):
    """Traduz uma celula de assinatura para {situacao, data}."""
    if valor is None:
        return {"situacao": "pendente", "data": None}

    texto = str(valor).strip()
    if not texto:
        return {"situacao": "pendente", "data": None}

    if sem_acento(texto).upper() in NAO_SE_APLICA:
        return {"situacao": "nao_se_aplica", "data": None}

    if PADRAO_ASSINADO.match(texto):
        achado = PADRAO_DATA.search(texto)
        return {"situacao": "assinado", "data": achado.group(1) if achado else None}

    return {"situacao": "pendente", "data": None}


def ler_data_envio(planilha):
    """Le a data do retrato ('DATA DE ENVIO: 17/08/2026') no topo da planilha."""
    for coluna in range(1, 6):
        valor = planilha.cell(LINHA_DATA_ENVIO, coluna).value
        if valor and "ENVIO" in sem_acento(str(valor)).upper():
            achado = PADRAO_DATA.search(str(valor))
            if achado:
                return datetime.strptime(achado.group(1), "%d/%m/%Y").replace(tzinfo=FUSO_LOCAL)
    erro(
        "Nao achei a 'DATA DE ENVIO' na linha %d da planilha.\n"
        "Ela e o que diz o quao atual e este retrato - sem ela eu nao consigo\n"
        "decidir quem manda quando a planilha e o e-mail discordarem." % LINHA_DATA_ENVIO
    )


def processar():
    caminho = achar_planilha()
    livro = openpyxl.load_workbook(caminho, data_only=True)
    if ABA not in livro.sheetnames:
        erro("A planilha nao tem a aba '%s'. Abas encontradas: %s" % (ABA, livro.sheetnames))
    planilha = livro[ABA]

    data_envio = ler_data_envio(planilha)

    nomes_signatarios = []
    for coluna in range(COL_PRIMEIRO_SIGNATARIO, COL_ULTIMO_SIGNATARIO + 1):
        bruto = planilha.cell(LINHA_CABECALHO_SIG, coluna).value
        if not bruto:
            erro("A coluna %d da linha %d esta sem nome de signatario." % (coluna, LINHA_CABECALHO_SIG))
        # "ASS. LUCAS G. QPC " -> "Lucas G. QPC"
        nome = re.sub(r"^ASS\.?\s*", "", str(bruto).replace("\n", " ")).strip()
        nome = re.sub(r"\s+", " ", nome)
        nomes_signatarios.append(nome.title().replace("Qpc", "QPC").replace("Lyon", "LYON").replace("Abr", "ABR"))

    contratos = []
    ignoradas = []
    for linha in range(PRIMEIRA_LINHA_DADOS, planilha.max_row + 1):
        identificacao = planilha.cell(linha, COL_IDENTIFICACAO).value
        if not identificacao:
            continue

        chave, e_aditivo = chave_do_documento(identificacao)
        if not chave:
            ignoradas.append((linha, str(identificacao)))
            continue

        signatarios = []
        for indice, coluna in enumerate(range(COL_PRIMEIRO_SIGNATARIO, COL_ULTIMO_SIGNATARIO + 1)):
            celula = ler_celula_assinatura(planilha.cell(linha, coluna).value)
            # Quem nao precisou assinar sai da lista inteira: nao aparece como
            # pendente, nao entra no denominador, nao vira cobranca indevida.
            if celula["situacao"] == "nao_se_aplica":
                continue
            signatarios.append({
                "quem": nomes_signatarios[indice],
                "papel": "signatario",
                "assinou": celula["situacao"] == "assinado",
                "data": celula["data"],
            })

        bruto_dias = planilha.cell(linha, COL_DIAS_SEM_ASSINATURA).value
        texto_dias = sem_acento(str(bruto_dias or "")).strip().upper()
        concluido = texto_dias.startswith("CONCLUID")
        dias_sem_assinatura = None
        if not concluido:
            achado = re.search(r"\d+", texto_dias)
            dias_sem_assinatura = int(achado.group(0)) if achado else None

        cadastro = planilha.cell(linha, COL_DATA_CADASTRO).value
        contratos.append({
            "chave": chave,
            "identificacao": re.sub(r"\s+", " ", str(identificacao)).strip(),
            "razao_social": (str(planilha.cell(linha, COL_RAZAO_SOCIAL).value or "").strip() or None),
            "servico": (str(planilha.cell(linha, COL_SERVICO).value or "").strip() or None),
            "cnpj": (str(planilha.cell(linha, COL_CNPJ).value or "").strip() or None),
            "data_cadastro": cadastro.strftime("%d/%m/%Y") if hasattr(cadastro, "strftime") else None,
            "aditivo": e_aditivo,
            "tipo": TIPO_DESTA_PLANILHA,
            "concluido": concluido,
            "dias_sem_assinatura": dias_sem_assinatura,
            "signatarios": signatarios,
            "linha_planilha": linha,
        })

    if not contratos:
        erro("A planilha nao rendeu nenhum contrato. Confira se o layout mudou.")

    # ---- Conferencia: chave repetida significa contrato duplicado -------------
    vistas = {}
    for contrato in contratos:
        if contrato["chave"] in vistas:
            erro(
                "Dois registros com a mesma etiqueta '%s':\n"
                "  linha %d: %s\n  linha %d: %s\n"
                "Isso quebra o cruzamento com os e-mails. Confira a planilha."
                % (contrato["chave"], vistas[contrato["chave"]]["linha_planilha"],
                   vistas[contrato["chave"]]["identificacao"],
                   contrato["linha_planilha"], contrato["identificacao"])
            )
        vistas[contrato["chave"]] = contrato

    saida = {
        "_leia_me": "Base consolidada lida da planilha F.SUP.G200.002. Gerada por 06-ler-planilha-controle.py. NAO EDITAR A MAO.",
        "fonte": os.path.basename(caminho),
        "data_do_retrato": data_envio.strftime("%Y-%m-%d"),
        "data_do_retrato_texto": data_envio.strftime("%d/%m/%Y"),
        "tipo": TIPO_DESTA_PLANILHA,
        "signatarios_do_fluxo": nomes_signatarios,
        "contratos": contratos,
    }
    with open(ARQUIVO_SAIDA, "w", encoding="utf-8") as arquivo:
        json.dump(saida, arquivo, ensure_ascii=False, indent=2)

    # ---- Mostra a conta, como manda a regra do projeto -----------------------
    so_contratos = [c for c in contratos if not c["aditivo"]]
    aditivos = [c for c in contratos if c["aditivo"]]
    concluidos = [c for c in contratos if c["concluido"]]
    print("=" * 70)
    print("BASE CONSOLIDADA - PLANILHA DE CONTROLE")
    print("=" * 70)
    print("Arquivo:            %s" % os.path.basename(caminho))
    print("Retrato de:         %s" % data_envio.strftime("%d/%m/%Y"))
    print("Linhas lidas:       %d  (%d contratos + %d aditivos)"
          % (len(contratos), len(so_contratos), len(aditivos)))
    if ignoradas:
        print("Linhas ignoradas:   %d (sem numero G200 reconhecivel)" % len(ignoradas))
        for linha, texto in ignoradas[:5]:
            print("    linha %d: %s" % (linha, texto[:60]))
    print("-" * 70)
    print("Concluidos:         %d" % len(concluidos))
    print("Em andamento:       %d" % (len(contratos) - len(concluidos)))
    pendentes = sum(
        len([s for s in c["signatarios"] if not s["assinou"]])
        for c in contratos if not c["concluido"]
    )
    print("Assinaturas faltando nos em andamento: %d" % pendentes)
    print("-" * 70)
    print("Fluxo de %d signatarios: %s" % (len(nomes_signatarios), ", ".join(nomes_signatarios)))
    print("=" * 70)
    print("Gravado em: %s" % ARQUIVO_SAIDA)


if __name__ == "__main__":
    processar()
